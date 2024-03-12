import openpyxl

widths = [9, 20, 10, 20]

def save_to_excel(users):
    # Создаем новую рабочую книгу Excel
    workbook = openpyxl.Workbook()

    # Получаем активный лист
    sheet = workbook.active

    # Заголовки столбцов
    headers = ["ID", "Имя", "Класс", "Пароль"]

    # Записываем заголовки в первую строку
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

        # устанавливаем ширину столбца
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = widths[col_num - 1]

    # Записываем данные в остальные строки
    for row_num, user in enumerate(users, 2):
        sheet.cell(row=row_num, column=1, value=user["id"])
        sheet.cell(row=row_num, column=2, value=user["name"])
        sheet.cell(row=row_num, column=3, value=user["grade"])
        sheet.cell(row=row_num, column=4, value=user["pass_key"])

    # Сохраняем рабочую книгу в файл
    excel_filename = "users_data.xlsx"
    workbook.save(excel_filename)

    return excel_filename